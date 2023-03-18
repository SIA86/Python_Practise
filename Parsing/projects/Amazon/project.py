from requests_html import HTMLSession
from random import choice
import time, sys, csv
from openpyxl import Workbook, load_workbook
from tkinter import *
from tkinter import messagebox
import tkinter.ttk as ttk

def main():
    url = ent_url.get()
    sub_url = 'https://www.amazon.com/s?'
    if not url:
        messagebox.showerror('Invalid URL!', "Parsing will not star while you don't enter result search URL")
    elif  sub_url not in url:
        messagebox.showerror('Invalid URL!', 'URL must begin with "https://www.amazon.com/s?"')
    elif all(x==0 for x in (s_name.get(), s_asin.get(), s_price.get(), s_link.get())):
        messagebox.showerror('Data type not chosen!', "Please choose at least one type of data for parsing")
    elif format.get() not in ['csv','xlsx']:
        messagebox.showerror('Invalid format!', "Please choose 'csv' or 'xlsx' output format")
    else:
        start = time.perf_counter()
        if_proxy = S_proxy.get()
        get_data(get_html(url, if_proxy), file_name(), if_proxy)
        print('Finish parsing!')
        fin = time.perf_counter() - start
        print('Total time: ', fin)
        messagebox.showinfo('Done!', f'Parsing has completed successfully. Total time {round(fin,2)} sec.')

# Collecting data from each page

def get_data(r, file_name, if_proxy):
    products = r.html.find('[data-component-type="s-search-result"]')
    print("Collecting product's data")

    for product in products:
        asin = product.attrs['data-asin']
        link = 'https://www.amazon.com' + product.find('span.rush-component a', first=True).attrs['href']

        name = product.find('span[class="a-size-base-plus a-color-base a-text-normal"]', first=True).text

        try:
            integer = product.find('span.a-price-whole', first=True).text
            decimal = product.find('span.a-price-fraction', first=True).text
            price = integer +  decimal
        except:
            price = None

        data = {}

        # Fields option
        if s_name.get() == 1:
            data['Name'] = name
        if s_asin.get() == 1:
            data['Asin'] = asin
        if s_price.get() == 1:
            data['Price'] = price
        if s_link.get() == 1:
            data['Link'] = link

        #Output format option
        if  format.get() == 'csv':
            csv_writer(data, file_name)
        elif format.get() == 'xlsx':
            xlsx_writer(data, file_name)

    #Getting next page URL
    try:
        next = r.html.find('a.s-pagination-next', first=True).attrs['href'] #Check if it is any next page
        get_data(get_html('https://www.amazon.com' + next, if_proxy), file_name, if_proxy)
    except AttributeError:
        pass

#  Enter search result page and try to get their HTML

def get_html(url, if_proxy):
    session = HTMLSession()
    headers = {'User_Agent':'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/107.0.0.0 Safari/537.36'}
    tries1 = 0
    tries2 = 0
    while(True):
        if if_proxy == 1: #Proxy option
            p = get_proxy()
            proxy = {p['schema']: p['adress']}
            print('Trying proxy:', proxy)
        else:
            proxy = None
        try:
            print('Connecting...')
            r = session.get(url, headers = headers, proxies=proxy)
        except Exception:
            print("Failed")
            print(f'{5-tries2} tries left')
            if tries2 < 5:
                time.sleep(2)
                tries2 += 1
                pass
            else:
                print("Couldn't establish connection")
                sys.exit(2)
        else:
            print('Connection established')
            if r.status_code == 200: # Check server response status
                r.html.render(sleep=0.5, timeout= 10, )
                return r
            else:
                print(f'Connection problems {r}')
                print(f'{5-tries1} tries left')
                if tries1 < 5:
                    time.sleep(5)
                    tries1 += 1
                    pass
                else:
                    print("Couldn't establish connection")
                    sys.exit(2)

#Collecting proxy-list from 'https://free-proxy-list.net/' and choose one random

def get_proxy():
    session = HTMLSession()
    try:
        r = session.get('https://free-proxy-list.net/')
    except Exception:
        print("Can't get proxies from https://free-proxy-list.net/")
        sys.exit()

    table = r.html.find('tbody tr')

    proxies = []

    for row in table[1:11]:
        ip = row.find('td')[0].text.strip()
        port = row.find('td')[1].text.strip()
        schema = 'https' if 'yes' in row.find('td')[6].text.strip() else 'http'
        proxy = {'schema':schema, 'adress': ip + ':' + port}
        proxies.append(proxy)

    return choice(proxies)

# Creating unique file name
def file_name():
    data = time.strftime("_%d_%b_%Y_")
    unique_number = str(round(time.time() * 100))
    return data+unique_number


# CSV file writing function
def csv_writer(data, file_name):
    print('Writing data to file')
    with open(f"Amazon{file_name}.csv", 'a', encoding='utf-8', newline='') as f:
        writer = csv.writer(f)
        writer.writerow(list(data.values()))


# xlsx file writing function
def xlsx_writer(data, file_name):
    print('Writing data to file')
    try:
        wb = load_workbook(filename = f"Amazon{file_name}.xlsx")
    except FileNotFoundError:
        wb = Workbook()

    finally:
        ws = wb.active
        ws.append(list(data.values()))
        wb.save(f"Amazon{file_name}.xlsx")


if __name__ == "__main__":
    
    #Create GUI

    root = Tk()
    root.title('Amazon Market Parser')
    root.resizable(width=False, height=False)

    style = ttk.Style()
    style.configure("Small.TLabel", font = ('Helvetica', 8))
    style.configure("Big.TLabel", font = ('Helvetica', 14))

    #Create parent frame - Frame 1
    frame1 = ttk.Frame(root, borderwidth=5, relief='sunken', padding='0 5 0 5')
    frame1.grid(column=0, row=0, sticky=W+E)

    #Create child frame - Frame 2
    frame2 = ttk.Frame(frame1)
    frame2.grid(column=0, row=0, sticky=W+E)

    ent_url = ttk.Entry(frame2, width=80)
    lbl_ent_url = ttk.Label(frame2, text='URL:', style = 'Big.TLabel')
    lbl_notation = ttk.Label(frame2, text='Note: choose the product what you are looking for and set all filter you need than copy the link \nfrom your browser to the field above', style='Small.TLabel')
    ttk.Separator(frame2).grid(row=2, columnspan=2, sticky=E+W, pady=5)


    ent_url.grid(column=1, row=0, padx=5)
    lbl_ent_url.grid(column=0, row=0, padx=5)
    lbl_notation.grid(column = 1, row=1, padx=5)

    #Create child frame - Frame 3
    frame3 = ttk.Frame(frame1)
    frame3.grid(column=0, row=1, sticky=W+E)

    #First column (filter options: name, price, asin, link)
    s_name = IntVar()
    s_price = IntVar()
    s_asin = IntVar()
    s_link = IntVar()
    s_name.set(0)
    s_price.set(0)
    s_asin.set(0)
    s_link.set(0)

    lbl_ch_filter = ttk.Label(frame3, text='Choose filter:')
    ch_name = ttk.Checkbutton(frame3, text='Name', variable=s_name)
    ch_price = ttk.Checkbutton(frame3, text='Price', variable=s_price)
    ch_asin = ttk.Checkbutton(frame3, text='Asin', variable=s_asin)
    ch_link = ttk.Checkbutton(frame3, text='Link', variable=s_link)

    lbl_ch_filter.grid(column=0, row=0, sticky=W, padx=5, pady=2)
    ch_name.grid(column=0, row=1, sticky=W, padx=5, pady=2)
    ch_price.grid(column=0, row=2, sticky=W, padx=5, pady=2)
    ch_asin.grid(column=0, row=3, sticky=W, padx=5, pady=2)
    ch_link.grid(column=0, row=4, sticky=W, padx=5, pady=2)

    #Second column (parser options: proxy)

    S_proxy = IntVar()
    S_proxy.set(0)

    lbl_parser_options = ttk.Label(frame3, text='Parser options:', padding='50 0 50 0')
    ch_proxy = ttk.Checkbutton(frame3, text='Use proxy', variable=S_proxy, padding='50 0 50 0')

    lbl_parser_options.grid(column=1, row=0,  padx=5, pady=2)
    ch_proxy.grid(column=1, row=1, sticky=W,  padx=5, pady=2)

    #Third column options (Output format: csv? xlsx)
    format = StringVar()

    lbl_output_options = ttk.Label(frame3, text='Output format:')
    rad_csv = ttk.Radiobutton(frame3, text='CSV', variable=format, value='csv')
    rad_xls = ttk.Radiobutton(frame3, text='XLSX', variable=format, value='xlsx')

    lbl_output_options.grid(column=2, row=0, sticky=W, padx=5, pady=2)
    rad_csv.grid(column=2, row=1, sticky=W, padx=5, pady=2)
    rad_xls.grid(column=2, row=2, sticky=W, padx=5, pady=2)

    #Create frame for buttons - Frame 4
    frame4 = ttk.Frame(root)
    frame4.grid(column=0, row=2, sticky=E)

    btn1 = ttk.Button(frame4, text='Start', command=main)
    btn2 = ttk.Button(frame4, text='Exit', command=root.quit)

    btn1.grid(column=0, row=0, padx=5, sticky=E)
    btn2.grid(column=1, row=0, padx=5, sticky=E)

    frame4.grid(column=0, row=2, sticky=E)

    root.mainloop()