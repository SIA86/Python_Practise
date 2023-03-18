from requests_html import HTMLSession
from multiprocessing import Pool
from random import choice
import time, sys, csv, json, math
from openpyxl import Workbook, load_workbook

import sqlite3

from tkinter import *
import tkinter.ttk as ttk



# Analizing main page code in order to find the link of the next page and the total amount of pages.
# Creating a list of URLS which later will be given as an argument to the parsing functions.


def pages(url):
    headers = {'User_Agent':'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/107.0.0.0 Safari/537.36'}
    session = HTMLSession()
    urls = [url]
    tries = 1
    while True:
        p = get_proxy() #Get the random proxy 
        proxy = {p['schema']: p['adress']}  
        print('Trying proxy: ', proxy) 
        try:
            r = session.get(url, headers=headers, proxies=proxy)
        except Exception:
            print('Failed')
            pass
        else:
            print('Success')
            if r.status_code == 200: #If get OK response begin to parse
                r.html.render(sleep=0.5, timeout = 20)
                script = r.html.xpath('////*[@id="search"]/script[9]', first=True).text[25:-2].replace("\\", "\\\\")
                #print(script)
                js = json.loads(script)          
                amount = math.ceil(js['totalResultCount']/24)
                print(str(amount) + ' pages found')
                if amount > 1:
                    for num in range(1, amount):
                        urls.append(url[:-6] + 'pg_' + str(num + 1)) #Create a list with URLS of all pages
                return urls
            else:
                print(f'Connection problems {r}') #If response is not 200 exit the program
                print(f'{5-tries} tries left')
                if tries <= 5:
                    time.sleep(5)
                    tries += 1
                    pass
                else:
                    sys.exit(2)
                


#  Enter each page from URLS list and try to get their HTML      
def get_html(url):
    #print(url)
    session = HTMLSession()
    headers = {'User_Agent':'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/107.0.0.0 Safari/537.36'}
    tries = 1    
    while(True):
        
        p = get_proxy()
        proxy = {p['schema']: p['adress']}
        print('Trying proxy: ', proxy)
        try:
            r = session.get(url, headers = headers, proxies=proxy)
        except Exception:
            print('Failed')
            pass
        else:
            print('Success')
            if r.status_code == 200:
                r.html.render( sleep=0.5, timeout= 15, )
                
                return r
            else:
                print(f'Connection problems {r}')
                print(f'{5-tries} tries left')
                if tries <= 5:
                    time.sleep(5)
                    tries += 1
                    pass
                else:
                    sys.exit(2)
                


# Collecting data from each page
def get_data(r):
        
    products = r.html.find('[data-component-type="s-search-result"]')  
    
    print("Collecting product's data")
 
    for product in products:
        asin = product.attrs['data-asin']
        link = 'https://www.amazon.com' + product.find('span.rush-component a', first=True).attrs['href']
                   
        name = product.find('span[class="a-size-base-plus a-color-base a-text-normal"]', first=True).text
        
        try: 
            integer = product.find('span.a-price-whole', first=True).text 
            decimal = product.find('span.a-price-fraction', first=True).text
            price = integer +  decimal
        except:
            price = None
        
        data = {"Asin": asin,
                "Name": name,
                "Price": price,
                "Link": link}
        
        #ADD try block if
        if  format.get() == 'csv':
            csv_writer(data)
        elif format.get() == 'xlsx':
            xlsx_writer(data)
        elif format.get() == 'db':
            db_writer(data)

        
    

        
        

# CSV file writing function
def csv_writer(data):
    print('Writing data to file')
    with open("Amazon2.csv", 'a', encoding='utf-8', newline='') as f:
        writer = csv.DictWriter(f, fieldnames=['Asin', 'Name', 'Price', 'Link'])
        writer.writerow({"Asin": data["Asin"],
                         "Name": data["Name"],
                         "Price": data['Price'],
                         "Link": data['Link']})


# xlsx file writing function
def xlsx_writer(data):
    print('Writing data to file')
    try:
        wb = load_workbook(filename = 'Amazon2.xlsx') 
    except FileNotFoundError:
        wb = Workbook()
        
    finally:
        ws = wb.active
        ws.append([data["Asin"], data["Name"], data['Price'], data['Link']]) 
        wb.save('Amazon2.xlsx') 


# db file writing function
def db_writer(data):
    print('Writing data to file')
    con = sqlite3.connect("Amazon2.db")
    cur = con.cursor()
    datas = (data["Asin"], data["Name"], data['Price'], data['Link'])
    try:
        cur.execute("CREATE TABLE result(Asin, Name, Price, Link)")
    except sqlite3.OperationalError:
        pass
    cur.execute("INSERT INTO result VALUES (?,?,?,?)", datas)
    con.commit()
    con.close()

    
# Grab a proxy from free-proxy web-site
def get_proxy():
    session = HTMLSession()
    r = session.get('https://free-proxy-list.net/')
    table = r.html.find('tbody tr')

    proxies = []

    for row in table[1:11]:
        ip = row.find('td')[0].text.strip()
        port = row.find('td')[1].text.strip()
        schema = 'https' if 'yes' in row.find('td')[6].text.strip() else 'http'
        proxy = {'schema':schema, 'adress': ip + ':' + port}
        proxies.append(proxy)

    return choice(proxies)

def make_all(url):
    get_data(get_html(url))


def main(): 
    
    base_url = ent_url.get()
    

    start = time.perf_counter()
    
    print('Getting pages urls')
    urls = pages(base_url)
    print('Begin parsing')
    #for url in urls:
    #    make_all(url, o_f)
    with Pool(2) as executor:
        executor.map(make_all, urls)
    print('Finish parsing!')
    
    fin = time.perf_counter() - start
    print('Total time: ', fin) 
    


#GUI interface
#GUI commands

def check_time_on():
    if S_timeout.get() == 1:
        ent_time.configure(state=NORMAL)
    else:
        ent_time.configure(state=DISABLED)

def check_multy_on():
    if S_multy.get() == 1:
        ent_multy.configure(state=NORMAL)
    else:
        ent_multy.configure(state=DISABLED)

#Create GUI
if __name__ == "__main__":
    root = Tk()
    root.title('Amazon Market Parser')
    root.resizable(width=False, height=False)

    #Create parent frame - Frame 1
    frame1 = ttk.Frame(root, borderwidth=5, relief='sunken', padding='0 5 0 5')
    frame1.grid(column=0, row=0, sticky=W+E)

    #Create child frame - Frame 2
    frame2 = ttk.Frame(frame1)
    frame2.grid(column=0, row=0)

    ent_url = ttk.Entry(frame2, width=80)
    lbl_ent_url = ttk.Label(frame2, text='URL:')
    lbl_notation = ttk.Label(frame2, text='Choose the product what you are looking for and set all filter you need \nthan copy the link from your browser to the field above')
    ttk.Separator(frame2).grid(row=2, columnspan=2, sticky=E+W, pady=5)

    ent_url.grid(column=1, row=0, padx=5)
    lbl_ent_url.grid(column=0, row=0, padx=5)
    lbl_notation.grid(column=1, row=1, padx=5)

    #Create child frame - Frame 3
    frame3 = ttk.Frame(frame1)
    frame3.grid(column=0, row=1, sticky=W+E)

    #First column (filter options: name, price, articul, marka)
    name = BooleanVar()
    price = BooleanVar()
    art = BooleanVar()
    mark = BooleanVar()
    name.set(0)
    price.set(0)
    art.set(0)
    mark.set(0)

    lbl_ch_filter = ttk.Label(frame3, text='Choose filter:')
    ch_name = ttk.Checkbutton(frame3, text='Name', variable=name)
    ch_price = ttk.Checkbutton(frame3, text='Price', variable=price)
    ch_art = ttk.Checkbutton(frame3, text='Art', variable=art)
    ch_mark = ttk.Checkbutton(frame3, text='Mark', variable=mark)

    lbl_ch_filter.grid(column=0, row=0, sticky=W, padx=5, pady=2)
    ch_name.grid(column=0, row=1, sticky=W, padx=5, pady=2)
    ch_price.grid(column=0, row=2, sticky=W, padx=5, pady=2)
    ch_art.grid(column=0, row=3, sticky=W, padx=5, pady=2)
    ch_mark.grid(column=0, row=4, sticky=W, padx=5, pady=2)

    #Second column (parser options)
    S_timeout = IntVar()
    S_proxy = IntVar()
    S_multy = IntVar()
    S_timeout.set(0)
    S_proxy.set(0)
    S_multy.set(0)

    lbl_parser_options = ttk.Label(frame3, text='Parser options:', padding='50 0 0 0')
    ch_proxy = ttk.Checkbutton(frame3, text='Use proxy', variable=S_proxy, padding='50 0 0 0')
    ch_time = ttk.Checkbutton(frame3, text='Timeout', variable=S_timeout, command=check_time_on, padding='50 0 0 0')
    ch_multy = ttk.Checkbutton(frame3, text='Multiprocessing', variable=S_multy, command=check_multy_on, padding='50 0 0 0')
    ent_time = ttk.Entry(frame3, width=5, state=DISABLED)
    ent_multy = ttk.Entry(frame3, width=5, state=DISABLED)

    lbl_parser_options.grid(column=1, row=0, sticky=W, padx=5, pady=2)
    ch_proxy.grid(column=1, row=1, sticky=W,  padx=5, pady=2)
    ch_time.grid(column=1, row=2, sticky=W,  padx=5, pady=2)
    ch_multy.grid(column=1, row=3, sticky=W,  padx=5, pady=2)
    ent_time.grid(column=2, row=2, sticky=W, padx=5, pady=2)
    ent_multy.grid(column=2, row=3, sticky=W,  padx=5, pady=2)

    #Third column options
    format = StringVar()

    lbl_output_options = ttk.Label(frame3, text='Output format:')
    rad_csv = ttk.Radiobutton(frame3, text='CSV', variable=format, value='csv')
    rad_db = ttk.Radiobutton(frame3, text='SQL', variable=format, value='db')
    rad_xls = ttk.Radiobutton(frame3, text='XLSX', variable=format, value='xlsx')

    lbl_output_options.grid(column=3, row=0, sticky=W, padx=5, pady=2)
    rad_csv.grid(column=3, row=1, sticky=W, padx=5, pady=2)
    rad_db.grid(column=3, row=2, sticky=W, padx=5, pady=2)
    rad_xls.grid(column=3, row=3, sticky=W, padx=5, pady=2)

    #Create frame for buttons - Frame 4
    frame4 = ttk.Frame(root)
    frame4.grid(column=0, row=2, sticky=E)

    btn1 = ttk.Button(frame4, text='Start', command=main)
    btn2 = ttk.Button(frame4, text='Exit', command=root.quit)

    btn1.grid(column=0, row=0, padx=5, sticky=E)
    btn2.grid(column=1, row=0, padx=5, sticky=E)

    frame4.grid(column=0, row=2, sticky=E)
    root.mainloop()
